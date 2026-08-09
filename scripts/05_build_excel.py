"""
05_build_excel.py
Builds Inventory_Demand_Forecasting_Model.xlsx -- the "Excel" half of the
resume line. All classification and inventory-policy math is done with real
Excel formulas (SUMIFS/RANK/INDEX-MATCH/IF), not pasted Python results, so
the workbook recalculates if inputs change. Demand statistics themselves
(avg/stddev daily demand) are sourced from the SQL layer (02_demand_stats.sql)
since they're computed off the 321k-row daily fact table -- that table is too
large to embed and recompute live in Excel, so it's cited as an external
source per SKU, exactly like a real analyst would pull a SQL extract into
Excel for the policy layer.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.comments import Comment

DATA = "/home/claude/project/data/"
OUT = "/home/claude/project/outputs/Inventory_Demand_Forecasting_Model.xlsx"

sku_master = pd.read_csv(DATA + "sku_master.csv")
demand_stats = pd.read_csv(DATA + "sql_out_demand_stats.csv")
onhand = pd.read_csv(DATA + "sku_onhand_network.csv")
model_comp = pd.read_csv(DATA + "forecast_model_comparison.csv")
next_wk_fc = pd.read_csv(DATA + "forecast_next_week.csv")
impact = pd.read_csv(DATA + "ai_before_after_impact.csv")
alerts = pd.read_csv(DATA + "ai_reorder_alerts.csv")

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT, color="0000FF", size=10)     # blue = hardcoded input
FORMULA_FONT = Font(name=FONT, color="000000", size=10)   # black = formula
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
SUB_FONT = Font(name=FONT, italic=True, size=10, color="595959")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ===========================================================================
# SHEET 0: README
# ===========================================================================
ws = wb.active
ws.title = "README"
ws["B2"] = "Inventory & Demand Forecasting Analytics for E-commerce Fulfillment"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Excel + SQL demand forecasting model, ABC analysis, and reorder-point / safety-stock calculator"
ws["B3"].font = SUB_FONT

readme_rows = [
    ("Sheet", "Contents"),
    ("SKU_Master", "SKU catalog: category, cost, price, supplier lead time, MOQ (input data)."),
    ("Demand_Stats", "90-day avg & std-dev daily demand per SKU, sourced from SQL (02_demand_stats.sql)."),
    ("ABC_Analysis", "Revenue-based Pareto classification (A/B/C) — LIVE Excel formulas (RANK, running SUM, IF)."),
    ("Reorder_Point_Calc", "Safety stock & reorder point per SKU — LIVE formulas, service level driven by ABC class."),
    ("Fill_Rate_Impact", "Baseline vs simulated fill rate after applying the new reorder policy (before/after)."),
    ("Forecast_Model", "ML demand-forecast model comparison (WMAPE) and next-week forecast per SKU."),
    ("Reorder_Alerts", "Automated, prioritized reorder alert list (urgency-ranked) — the AI-automation output."),
]
r = 5
for a, b in readme_rows:
    ws.cell(row=r, column=2, value=a).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=3, value=b).font = Font(name=FONT, size=10)
    r += 1

ws["B15"] = "Color legend:"
ws["B15"].font = Font(name=FONT, bold=True, size=10)
ws["B16"] = "Blue text = hardcoded input / external source"
ws["B16"].font = INPUT_FONT
ws["B17"] = "Black text = live Excel formula"
ws["B17"].font = FORMULA_FONT
autosize(ws, [3, 22, 90])

# ===========================================================================
# SHEET 1: SKU_Master
# ===========================================================================
ws = wb.create_sheet("SKU_Master")
cols = ["sku_id", "sku_name", "category", "unit_cost", "unit_price",
        "supplier_lead_time_days", "moq_units", "assigned_fcs"]
headers = ["SKU ID", "SKU Name", "Category", "Unit Cost (₹)", "Unit Price (₹)",
           "Lead Time (days)", "MOQ (units)", "Fulfillment Centers"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))
for i, row in sku_master.iterrows():
    for c, col in enumerate(cols, start=1):
        cell = ws.cell(row=i + 2, column=c, value=row[col])
        cell.font = INPUT_FONT
        cell.border = BORDER
        if col in ("unit_cost", "unit_price"):
            cell.number_format = "₹#,##0"
ws.freeze_panes = "A2"
autosize(ws, [10, 22, 24, 13, 13, 14, 12, 26])
n_sku = len(sku_master)

# ===========================================================================
# SHEET 2: Demand_Stats  (sourced from SQL layer)
# ===========================================================================
ws = wb.create_sheet("Demand_Stats")
headers = ["SKU ID", "Days Observed", "Avg Daily Demand", "Std Dev Daily Demand"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))
demand_stats_sorted = demand_stats.set_index("sku_id").reindex(sku_master["sku_id"]).reset_index()
for i, row in demand_stats_sorted.iterrows():
    ws.cell(row=i + 2, column=1, value=row["sku_id"]).font = INPUT_FONT
    ws.cell(row=i + 2, column=2, value=int(row["days_observed"])).font = INPUT_FONT
    ws.cell(row=i + 2, column=3, value=float(row["avg_daily_demand"])).font = INPUT_FONT
    ws.cell(row=i + 2, column=4, value=float(row["stddev_daily_demand"])).font = INPUT_FONT
    for c in range(1, 5):
        ws.cell(row=i + 2, column=c).border = BORDER
ws["F1"] = "Source: SQL view v_sku_demand_stats_90d (02_demand_stats.sql), trailing 90 days of daily_sales fact table."
ws["F1"].font = SUB_FONT
ws.freeze_panes = "A2"
autosize(ws, [10, 14, 16, 18])

# ===========================================================================
# SHEET 3: ABC_Analysis (LIVE FORMULAS)
# ===========================================================================
ws = wb.create_sheet("ABC_Analysis")
headers = ["SKU ID", "Category", "Units Sold (12M)", "Revenue (12M, ₹)",
           "Revenue Rank", "Cumulative Revenue %", "ABC Class"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))

abc_src = pd.read_csv(DATA + "sql_out_abc_classification.csv").set_index("sku_id").reindex(sku_master["sku_id"]).reset_index()
# write inputs (units, revenue) as values (sourced from SQL) -- rank/cum%/class computed live in Excel
for i, row in abc_src.iterrows():
    r = i + 2
    ws.cell(row=r, column=1, value=row["sku_id"]).font = INPUT_FONT
    ws.cell(row=r, column=2, value=row["category"]).font = INPUT_FONT
    ws.cell(row=r, column=3, value=int(row["units_sold_12m"])).font = INPUT_FONT
    rev_cell = ws.cell(row=r, column=4, value=float(row["revenue_12m"]))
    rev_cell.font = INPUT_FONT
    rev_cell.number_format = "₹#,##0"

last_row = n_sku + 1
for i in range(n_sku):
    r = i + 2
    # Revenue rank (live formula)
    rank_cell = ws.cell(row=r, column=5, value=f"=RANK(D{r},$D$2:$D${last_row})")
    rank_cell.font = FORMULA_FONT
    # Cumulative % of total revenue for SKUs ranked <= this SKU's rank (live formula)
    cum_cell = ws.cell(row=r, column=6,
                        value=f"=SUMPRODUCT(($E$2:$E${last_row}<=E{r})*$D$2:$D${last_row})/SUM($D$2:$D${last_row})")
    cum_cell.number_format = "0.0%"
    cum_cell.font = FORMULA_FONT
    # ABC class (live formula, standard 80/95 Pareto cut)
    class_cell = ws.cell(row=r, column=7, value=f'=IF(F{r}<=0.8,"A",IF(F{r}<=0.95,"B","C"))')
    class_cell.font = FORMULA_FONT
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER

# conditional formatting to color the ABC class column
ws.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"A"'], fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"B"'], fill=PatternFill("solid", fgColor="FFEB9C")))
ws.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"C"'], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.freeze_panes = "A2"
autosize(ws, [10, 24, 16, 16, 12, 18, 10])

# ABC summary block
ws["I1"] = "ABC Summary"
ws["I1"].font = Font(name=FONT, bold=True, size=11)
ws["I2"] = "Class"; ws["J2"] = "SKU Count"; ws["K2"] = "% of SKUs"; ws["L2"] = "% of Revenue"
for c in ["I2", "J2", "K2", "L2"]:
    ws[c].font = Font(name=FONT, bold=True, size=10)
for idx, cls in enumerate(["A", "B", "C"]):
    r = 3 + idx
    ws.cell(row=r, column=9, value=cls).font = FORMULA_FONT
    ws.cell(row=r, column=10, value=f'=COUNTIF($G$2:$G${last_row},I{r})').font = FORMULA_FONT
    pc = ws.cell(row=r, column=11, value=f'=J{r}/{n_sku}')
    pc.number_format = "0.0%"; pc.font = FORMULA_FONT
    pr = ws.cell(row=r, column=12, value=f'=SUMIF($G$2:$G${last_row},I{r},$D$2:$D${last_row})/SUM($D$2:$D${last_row})')
    pr.number_format = "0.0%"; pr.font = FORMULA_FONT

# Pie chart of revenue share by class
pie = PieChart()
pie.title = "Revenue Contribution by ABC Class"
data = Reference(ws, min_col=12, min_row=2, max_row=5)
cats = Reference(ws, min_col=9, min_row=3, max_row=5)
pie.add_data(data, titles_from_data=False)
pie.set_categories(cats)
pie.height = 7.5; pie.width = 11
ws.add_chart(pie, "I8")

# ===========================================================================
# SHEET 4: Reorder_Point_Calc (LIVE FORMULAS, cross-sheet)
# ===========================================================================
ws = wb.create_sheet("Reorder_Point_Calc")
headers = ["SKU ID", "ABC Class", "Avg Daily Demand", "Std Dev Demand", "Lead Time (days)",
           "Service Level Z", "Safety Stock (units)", "Reorder Point (units)",
           "Net On-Hand (units)", "Action"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header_row(ws, 1, len(headers))

onhand_sorted = onhand.set_index("sku_id").reindex(sku_master["sku_id"]).reset_index().fillna(0)

for i in range(n_sku):
    r = i + 2
    sku_id = sku_master.iloc[i]["sku_id"]
    ws.cell(row=r, column=1, value=f"=SKU_Master!A{r}").font = FORMULA_FONT
    ws.cell(row=r, column=2, value=f"=ABC_Analysis!G{r}").font = FORMULA_FONT
    ws.cell(row=r, column=3, value=f"=Demand_Stats!C{r}").font = FORMULA_FONT
    ws.cell(row=r, column=4, value=f"=Demand_Stats!D{r}").font = FORMULA_FONT
    ws.cell(row=r, column=5, value=f"=SKU_Master!F{r}").font = FORMULA_FONT
    # service level Z by ABC class: A=98% (2.05), B=95% (1.65), C=90% (1.28)
    z_cell = ws.cell(row=r, column=6, value=f'=IF(B{r}="A",2.05,IF(B{r}="B",1.65,1.28))')
    z_cell.font = FORMULA_FONT
    ss_cell = ws.cell(row=r, column=7, value=f"=ROUND(F{r}*D{r}*SQRT(E{r}),1)")
    ss_cell.font = FORMULA_FONT
    rop_cell = ws.cell(row=r, column=8, value=f"=ROUND((C{r}*E{r})+G{r},1)")
    rop_cell.font = FORMULA_FONT
    onh_cell = ws.cell(row=r, column=9, value=float(onhand_sorted.iloc[i]["on_hand_units"]))
    onh_cell.font = INPUT_FONT
    act_cell = ws.cell(row=r, column=10, value=f'=IF(I{r}<=H{r},"REORDER NOW","OK")')
    act_cell.font = FORMULA_FONT
    for c in range(1, 11):
        ws.cell(row=r, column=c).border = BORDER

ws.conditional_formatting.add(f"J2:J{last_row}",
    CellIsRule(operator="equal", formula=['"REORDER NOW"'], fill=PatternFill("solid", fgColor="FFC7CE")))
ws["L1"] = "Formulas: Safety Stock = Z × σ(demand) × √(Lead Time)  |  Reorder Point = (Avg Demand × Lead Time) + Safety Stock"
ws["L1"].font = SUB_FONT
ws["L2"] = "Net On-Hand: Source = current inventory snapshot (SQL, inventory_snapshot table), network-wide sum across assigned FCs."
ws["L2"].font = SUB_FONT
ws.freeze_panes = "A2"
autosize(ws, [10, 11, 15, 14, 13, 13, 16, 17, 15, 13])

# ===========================================================================
# SHEET 5: Fill_Rate_Impact
# ===========================================================================
ws = wb.create_sheet("Fill_Rate_Impact")
ws["B2"] = "Before vs After: Reorder Policy Impact (Last 12 Weeks)"
ws["B2"].font = Font(name=FONT, bold=True, size=12)
ws["B3"] = "AFTER = simulated result of replacing the naive weekly-review policy with the ABC-tiered, forecast-driven reorder-point policy from this workbook."
ws["B3"].font = SUB_FONT

metrics = [
    ("Fill Rate — BEFORE (naive weekly-review policy)", float(impact["fill_rate_before_pct"][0]), "0.0%"),
    ("Fill Rate — AFTER (ABC-tiered, forecast-driven policy)", float(impact["fill_rate_after_pct"][0]), "0.0%"),
    ("Improvement (percentage points)", float(impact["improvement_pp"][0]), "0.0"),
    ("Total Demand, units (12-wk window)", int(impact["demand_units"][0]), "#,##0"),
    ("Lost Sales — BEFORE, units", int(impact["lost_sales_before"][0]), "#,##0"),
    ("Lost Sales — AFTER (estimated), units", int(impact["lost_sales_after_est"][0]), "#,##0"),
]
r = 5
for label, val, fmt in metrics:
    ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10)
    v = val / 100 if "%" in fmt else val
    vc = ws.cell(row=r, column=4, value=v)
    vc.number_format = fmt
    vc.font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    r += 1

bar = BarChart()
bar.title = "Fill Rate: Before vs After"
bar.y_axis.title = "Fill Rate"
bar.type = "col"
data = Reference(ws, min_col=4, min_row=5, max_row=6)
cats = Reference(ws, min_col=2, min_row=5, max_row=6)
bar.add_data(data, titles_from_data=False)
bar.set_categories(cats)
bar.height = 7.5; bar.width = 13
ws.add_chart(bar, "B13")
autosize(ws, [3, 50, 3, 16])

# ===========================================================================
# SHEET 6: Forecast_Model
# ===========================================================================
ws = wb.create_sheet("Forecast_Model")
ws["B2"] = "Demand Forecasting Model — Benchmark (WMAPE %, lower = better)"
ws["B2"].font = Font(name=FONT, bold=True, size=12)
headers = ["Model", "WMAPE %"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=4, column=c + 1, value=h)
style_header_row(ws, 4, 2)
model_comp_sorted = model_comp.sort_values("wmape_pct")
for i, row in model_comp_sorted.iterrows():
    r = 5 + i
    ws.cell(row=r, column=2, value=row["model"]).font = INPUT_FONT
    v = ws.cell(row=r, column=3, value=float(row["wmape_pct"]))
    v.number_format = "0.00"
    v.font = INPUT_FONT
    for c in [2, 3]:
        ws.cell(row=r, column=c).border = BORDER
ws["B10"] = "Winning model: Random Forest Regressor (engineered lag + rolling-window + calendar features, weekly SKU-level demand)."
ws["B10"].font = SUB_FONT
ws["B11"] = "Trained/evaluated in Python (scikit-learn); see 03_forecasting_model.py. Time-based 8-week holdout, no leakage (all features causal)."
ws["B11"].font = SUB_FONT

ws["B14"] = "Next-Week Demand Forecast (sample, top 15 by forecast volume)"
ws["B14"].font = Font(name=FONT, bold=True, size=11)
headers2 = ["SKU ID", "Category", "Forecast Week", "Forecast Demand (units)"]
for c, h in enumerate(headers2, start=1):
    ws.cell(row=15, column=c + 1, value=h)
style_header_row(ws, 15, 4)
top15 = next_wk_fc.sort_values("forecast_demand_units", ascending=False).head(15).reset_index(drop=True)
for i, row in top15.iterrows():
    r = 16 + i
    ws.cell(row=r, column=2, value=row["sku_id"]).font = INPUT_FONT
    ws.cell(row=r, column=3, value=row["category"]).font = INPUT_FONT
    ws.cell(row=r, column=4, value=str(row["forecast_week"])).font = INPUT_FONT
    ws.cell(row=r, column=5, value=float(row["forecast_demand_units"])).font = INPUT_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = BORDER
autosize(ws, [3, 26, 24, 16, 22])

# ===========================================================================
# SHEET 7: Reorder_Alerts (AI automation output)
# ===========================================================================
ws = wb.create_sheet("Reorder_Alerts")
ws["B1"] = "Automated Reorder Alerts (AI Automation Layer — prioritized, urgency-ranked)"
ws["B1"].font = Font(name=FONT, bold=True, size=12)
headers = ["SKU ID", "SKU Name", "Category", "ABC Class", "On Hand", "Days of Cover",
           "Lead Time (days)", "Forecast Demand (next wk)", "Urgency", "Suggested Order Qty"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=3, column=c, value=h)
style_header_row(ws, 3, len(headers))
alert_view = alerts[alerts["needs_reorder"] == True].sort_values(
    by=["urgency"], key=lambda s: s.map({"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2}))
for i, row in alert_view.iterrows():
    r = 4 + list(alert_view.index).index(i)
    vals = [row["sku_id"], row["sku_name"], row["category"], row["abc_class"],
            row["on_hand_units"], row["days_of_cover"], row["supplier_lead_time_days"],
            row["forecast_demand_units"], row["urgency"], row["suggested_order_qty"]]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = INPUT_FONT
        cell.border = BORDER
last_alert_row = 3 + len(alert_view)
ws.conditional_formatting.add(f"I4:I{last_alert_row}",
    CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill("solid", fgColor="FF9999")))
ws.conditional_formatting.add(f"I4:I{last_alert_row}",
    CellIsRule(operator="equal", formula=['"HIGH"'], fill=PatternFill("solid", fgColor="FFD699")))
ws.conditional_formatting.add(f"I4:I{last_alert_row}",
    CellIsRule(operator="equal", formula=['"MEDIUM"'], fill=PatternFill("solid", fgColor="FFF2AE")))
ws.freeze_panes = "A4"
autosize(ws, [10, 22, 24, 10, 10, 13, 13, 20, 11, 17])

for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print(f"Workbook saved: {OUT}")
